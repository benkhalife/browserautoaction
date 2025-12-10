import openpyxl

excel_users = "UsersExam.xlsx"

excel_support = "./exams/support.xlsx"

try:
    workbook = openpyxl.load_workbook(excel_users)
    workbook_support = openpyxl.load_workbook(excel_support)

    sheet = workbook.active
    sheet_support = workbook_support.active

    index = 0
    for user in sheet.iter_rows(values_only=True):
        index += 1

        if index == 1:
            continue
        (email, password, exam_user_link, staus) = user

        index_support = 0
        for exam in sheet_support.iter_rows(values_only=True):
            index_support += 1
            if index_support == 1:
                continue
            (exam_file, exam_path, exam_link) = exam
            if exam_link == exam_user_link:
                exam_excel_file = f"{exam_path}/{exam_file}"
                print(exam_excel_file)

    # # استخراج مقادیر خانه‌های A1 و A2
    # login_url = sheet["A1"].value
    # exam_url = sheet["B1"].value

    # # نمایش نتایج برای اطمینان (اختیاری)
    # print(f"login_url: {login_url}\n")
    # print(f"exam_url: {exam_url}")

    # کل داده‌های اکسل در متغیر workbook ذخیره شده است
    # می‌توانید از workbook برای دسترسی به سایر داده‌ها استفاده کنید

except FileNotFoundError:
    print(f"خطا: فایل '{excel_users}' یافت نشد. لطفاً نام فایل و مسیر را بررسی کنید.")
except Exception as e:
    print(f"خطای غیرمنتظره رخ داد: {str(e)}")
